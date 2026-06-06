import io
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def generate_excel_from_dict_list(data: list[dict[str, Any]], sheet_name: str = "Report") -> io.BytesIO:
    """
    Generates an Excel file from a list of dictionaries.
    Uses the keys of the first dictionary as headers.
    Returns a BytesIO object suitable for a StreamingResponse.
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limits sheet name to 31 chars

    if not data:
        wb.save(output)
        output.seek(0)
        return output

    # 1. Extract headers from the first dictionary
    headers = list(data[0].keys())
    
    # 2. Write Headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(header).replace("_", " ").title())
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 3. Write Data
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            val = row_data.get(header)
            # Basic formatting
            if isinstance(val, (int, float)):
                ws.cell(row=row_idx, column=col_idx, value=val)
            elif val is None:
                ws.cell(row=row_idx, column=col_idx, value="")
            else:
                ws.cell(row=row_idx, column=col_idx, value=str(val))

    # 4. Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_length = len(str(header))
        for row in range(2, len(data) + 2):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50) # Cap at 50

    wb.save(output)
    output.seek(0)
    return output
